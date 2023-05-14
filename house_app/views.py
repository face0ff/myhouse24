from django.contrib import messages
from django.contrib.auth.decorators import user_passes_test
from django.core.mail import EmailMessage
from django.db import IntegrityError
from django.db.models import Q, Sum
from django.forms import model_to_dict
from django.http import JsonResponse, HttpResponseRedirect, HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.template.loader import render_to_string
from django.urls import reverse_lazy
from django.utils.decorators import method_decorator
from django.views.generic import DetailView, ListView, CreateView, UpdateView
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from house_app.forms import *
from house_app.models import House, Apartment, Section, Floor, Request
from services_app.models import MeterReading, PriceTariffServices
from user_app.models import UserProfile, Role
from openpyxl import load_workbook
import locale
from openpyxl.styles import Font, Alignment, Side, Border
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle



def check_user_is_staff(user, obj):
    if hasattr(user.role, obj) and getattr(user.role, obj) or user.is_superuser:
        return True
    return False

# Create your views here.
@method_decorator(user_passes_test(lambda u: check_user_is_staff(u, 'house'), login_url='login_admin'), name='dispatch')
class HousesList(ListView):
    model = House
    template_name = 'houses_list.html'

    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(**kwargs)
        object_list = House.objects.all()
        if self.request.GET.get('name'):
            object_list = object_list.filter(name__contains=self.request.GET.get('name'))
        if self.request.GET.get('address'):
            object_list = object_list.filter(address__contains=self.request.GET.get('address'))
        if self.request.GET.get('filter-name') == '1':
            object_list = object_list.order_by('-name')
        if self.request.GET.get('filter-name') == '0':
            object_list = object_list.order_by('name')
        if self.request.GET.get('filter-address') == '1':
            object_list = object_list.order_by('-address')
        if self.request.GET.get('filter-address') == '0':
            object_list = object_list.order_by('address')

        context['object_list'] = object_list
        return context


class HouseDetail(DetailView):
    model = House
    template_name = 'house_detail.html'


class HouseCreate(CreateView):
    model = House
    template_name = 'house_create.html'
    form_class = HouseForm
    success_url = reverse_lazy('houses_list')

    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(**kwargs)
        context['section_formset'] = SectionFormSet(queryset=Section.objects.none(), prefix='section')
        context['floor_formset'] = FloorFormSet(queryset=Floor.objects.none(), prefix='floor')
        context['user_formset'] = UserFormSet(prefix='user')
        context['house_form'] = HouseForm()
        return context

    def post(self, *args, **kwargs):
        print(self.request.POST)
        self.object = None
        house_form = HouseForm(self.request.POST, self.request.FILES)
        section_formset = SectionFormSet(self.request.POST, prefix='section')
        floor_formset = FloorFormSet(self.request.POST, prefix='floor')
        user_formset = UserFormSet(self.request.POST, prefix='user')

        if all([section_formset.is_valid(), floor_formset.is_valid(), user_formset.is_valid(), house_form.is_valid()]):
            return self.form_valid(section_formset, floor_formset, user_formset, house_form)
        else:
            return self.form_invalid(section_formset, floor_formset, user_formset, house_form)

    def form_valid(self, section_formset, floor_formset, user_formset, house_form):

        house = house_form.save(commit=False)
        house.save()
        section_formset.save(commit=False)
        floor_formset.save(commit=False)

        for form in section_formset:
            if form.is_valid():
                item = form.save(commit=False)
                item.house_id = house.id
                try:
                    item.save()
                except IntegrityError:
                    pass
            else:
                return self.form_invalid(section_formset, floor_formset, user_formset, house_form)
        for form in section_formset.deleted_objects:
            form.delete()
        messages.success(self.request, "Valid form")

        for form in floor_formset:
            if form.is_valid():
                item = form.save(commit=False)
                item.house_id = house.id
                try:
                    item.save()
                except IntegrityError:
                    pass
            else:
                return self.form_invalid(section_formset, floor_formset, user_formset, house_form)
        for form in floor_formset.deleted_objects:
            form.delete()
        messages.success(self.request, "Valid form")

        for form_user in user_formset:
            if form_user.cleaned_data and form_user.cleaned_data['DELETE'] is False:
                user = form_user.cleaned_data.get('user')
                house.users.add(user)

        # for form in user_formset.deleted_objects:
        #     form.delete()
        messages.success(self.request, "Valid form")
        return HttpResponseRedirect(self.success_url)

    def form_invalid(self, section_formset, floor_formset, user_formset, house_form, **kwargs):
        print(section_formset.errors)
        print(floor_formset.errors)
        print(user_formset.errors)
        print(house_form.errors)
        messages.error(self.request, "Invalid form")
        return self.render_to_response(
            self.get_context_data(section_formset=section_formset, floor_formset=floor_formset,
                                  user_formset=user_formset, house_form=house_form))


def select_user(request):
    if request.GET.get('user_id'):
        print(request.GET.get('user_id'))
        user_role = UserProfile.objects.get(pk=request.GET.get('user_id')).role.get_roles_display()
        role_id = UserProfile.objects.get(pk=request.GET.get('user_id')).role_id
        print(user_role)
        data = {
            'user_role': user_role,
            'role_id': role_id
        }
        return JsonResponse(data, status=200)


class HouseUpdate(UpdateView):
    model = House
    template_name = 'house_update.html'
    form_class = HouseForm
    success_url = reverse_lazy('houses_list')

    # queryset = House.objects.prefetch_related('users__role')

    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(**kwargs)
        context['section_formset'] = SectionFormSet(queryset=Section.objects.filter(house_id=self.get_object().pk),
                                                    prefix='section')
        context['floor_formset'] = FloorFormSet(queryset=Floor.objects.filter(house_id=self.get_object().pk),
                                                prefix='floor')
        users = self.object.users.all()
        context['user_formset'] = UserFormSet(prefix='user', initial=[{'user': item.id,
                                                                       'role': item.role.get_roles_display()} for item
                                                                      in users])
        context['house_form'] = HouseForm(instance=get_object_or_404(House, id=self.get_object().pk))
        return context

    def post(self, *args, **kwargs):
        print(self.request.POST)
        self.object = self.get_object()
        house_form = HouseForm(self.request.POST, self.request.FILES,
                               instance=get_object_or_404(House, id=self.get_object().pk))
        section_formset = SectionFormSet(self.request.POST, prefix='section',
                                         queryset=Section.objects.filter(house_id=self.get_object().pk))
        floor_formset = FloorFormSet(self.request.POST, prefix='floor',
                                     queryset=Floor.objects.filter(house_id=self.get_object().pk))
        user_formset = UserFormSet(self.request.POST, prefix='user')

        if all([section_formset.is_valid(), floor_formset.is_valid(), user_formset.is_valid(), house_form.is_valid()]):
            return self.form_valid(section_formset, floor_formset, user_formset, house_form)
        else:
            return self.form_invalid(section_formset, floor_formset, user_formset, house_form)

    def form_valid(self, section_formset, floor_formset, user_formset, house_form):

        house = house_form.save(commit=False)
        house.save()
        section_formset.save(commit=False)
        floor_formset.save(commit=False)

        for form in section_formset:
            if form.is_valid():
                item = form.save(commit=False)
                item.house_id = house.id
                try:
                    item.save()
                except IntegrityError:
                    pass
            else:
                return self.form_invalid(section_formset, floor_formset, user_formset, house_form)
        for form in section_formset.deleted_objects:
            form.delete()
        messages.success(self.request, "Valid form")

        for form in floor_formset:
            if form.is_valid():
                item = form.save(commit=False)
                item.house_id = house.id
                try:
                    item.save()
                except IntegrityError:
                    pass
            else:
                return self.form_invalid(section_formset, floor_formset, user_formset, house_form)
        for form in floor_formset.deleted_objects:
            form.delete()
        messages.success(self.request, "Valid form")
        house.users.clear()
        for form_user in user_formset:
            if form_user.cleaned_data and form_user.cleaned_data['DELETE'] is False:
                user = form_user.cleaned_data.get('user')
                house.users.add(user)

        messages.success(self.request, "Valid form")
        return HttpResponseRedirect(self.success_url)

    def form_invalid(self, section_formset, floor_formset, user_formset, house_form, **kwargs):
        print(section_formset.errors)
        print(floor_formset.errors)
        print(user_formset.errors)
        print(house_form.errors)
        messages.error(self.request, "Invalid form")
        return self.render_to_response(
            self.get_context_data(section_formset=section_formset, floor_formset=floor_formset,
                                  user_formset=user_formset, house_form=house_form))


def house_delete(request, pk):
    try:
        item = get_object_or_404(House, id=pk)
        item.delete()
    except:
        print('neviydet')
    return redirect('houses_list')


class ApartmentDetail(DetailView):
    model = Apartment
    template_name = 'apartment_detail.html'

@method_decorator(user_passes_test(lambda u: check_user_is_staff(u, 'apartment'), login_url='login_admin'), name='dispatch')
class ApartmentsList(ListView):
    model = Apartment
    template_name = 'apartments_list.html'
    context_object_name = 'apartments'

    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(**kwargs)
        houses = House.objects.all()
        queryset = Apartment.objects.all()

        # personal_accounts = PersonalAccount.objects.select_related('apartment').all()
        # context['personal_accounts'] = personal_accounts
        # context['idx'] = [x.apartment.id for x in personal_accounts if x.apartment]

        if self.request.GET.get('filter-number') == '1':
            queryset = queryset.order_by('-number')
        if self.request.GET.get('filter-number') == '0':
            queryset = queryset.order_by('number')
        if self.request.GET.get('filter-house') == '1':
            queryset = queryset.order_by('-id')
        if self.request.GET.get('filter-house') == '0':
            queryset = queryset.order_by('id')
        if self.request.GET.get('filter-section') == '1':
            queryset = queryset.order_by('-section')
        if self.request.GET.get('filter-section') == '0':
            queryset = queryset.order_by('section')
        if self.request.GET.get('filter-floor') == '1':
            queryset = queryset.order_by('-floor')
        if self.request.GET.get('filter-floor') == '0':
            queryset = queryset.order_by('floor')
        if self.request.GET.get('filter-owner') == '0':
            queryset = queryset.order_by('-owner__last_name')
        if self.request.GET.get('filter-owner') == '1':
            queryset = queryset.order_by('owner__last_name')
        if self.request.GET.get('input_number'):
            queryset = queryset.filter(number__icontains=self.request.GET.get('input_number'))
        if self.request.GET.get('input_house'):
            queryset = queryset.filter(house_id=self.request.GET.get('input_house'))
        if self.request.GET.get('input_section'):
            queryset = queryset.filter(section__name=self.request.GET.get('input_section'))
        if self.request.GET.get('input_floor'):
            queryset = queryset.filter(floor__name=self.request.GET.get('input_floor'))
        if self.request.GET.get('input_owner'):
            queryset = queryset.filter(owner=self.request.GET.get('input_owner'))
        # if self.request.GET.get('debt'):
        #     if self.request.GET.get('debt') == 'yes':
        #         p_a = PersonalAccount.objects.select_related('apartment').filter(balance__lt=0)
        #         idx = [x.apartment.id for x in p_a if x.apartment]
        #     else:
        #         p_a = PersonalAccount.objects.select_related('apartment').filter(balance__gte=0)
        #         idx = [x.apartment.id for x in p_a if x.apartment]
        #     queryset = queryset.filter(pk__in=idx)
        if self.request.GET.get('input_house'):
            for house in houses:
                if house.id == int(self.request.GET.get('input_house')):
                    context['sections'] = Section.objects.filter(house=house.id)
                    context['floors'] = Floor.objects.filter(house=house.id)

        paginator = Paginator(queryset, 10)  # 3 posts in each page
        page = self.request.GET.get('page')
        try:
            queryset = paginator.page(page)
        except PageNotAnInteger:
            # If page is not an integer deliver the first page
            queryset = paginator.page(1)
        except EmptyPage:
            # If page is out of range deliver last page of results
            queryset = paginator.page(paginator.num_pages)

        context['page'] = page
        context['houses'] = houses
        context['owners'] = UserProfile.objects.filter(is_staff=False)
        context['apartments'] = queryset
        return context


class ApartmentCreate(CreateView):
    model = Apartment
    template_name = 'apartment_create.html'
    form_class = ApartmentForm
    success_url = reverse_lazy('apartments_list')

    # def form_valid(self, form):
    #     super(ApartmentCreate, self).form_valid(form)
    #     return redirect('apartment_detail', self.object.id)

    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(**kwargs)
        queryset = Apartment.objects.all()

        context['accounts'] = Account.objects.filter(apartment__isnull=True)
        context['apartments'] = queryset
        return context


class ApartmentUpdate(UpdateView):
    model = Apartment
    template_name = 'apartment_update.html'
    form_class = ApartmentForm
    success_url = reverse_lazy('apartments_list')

    def form_valid(self, form):
        if 'action_save_add' in self.request.POST:
            # Получить последний идентификатор
            last_id = Apartment.objects.all().order_by('-id').first().id
            last_number = Apartment.objects.all().order_by('-id').first().number
            # Увеличить на единицу
            new_id = last_id + 1
            # Установить имя
            new_number = last_number + 1
            # Сохранить новые значения
            form.instance.id = new_id
            form.instance.number = new_number
            super(ApartmentUpdate, self).form_valid(form)
            return redirect('apartment_update', new_id)
        else:
            super(ApartmentUpdate, self).form_valid(form)
            return redirect('apartment_detail', self.object.id)

    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(**kwargs)
        queryset = Apartment.objects.all()

        context['accounts'] = Account.objects.filter(apartment__isnull=True)
        context['apartments'] = queryset
        return context


def apartment_delete(request, pk):
    try:
        apart = get_object_or_404(Apartment, id=pk)
        apart.delete()
    except:
        print('nelzia')
    return redirect('apartments_list')


def select_house(request):
    if request.GET.get('house_field'):
        house = House.objects.get(pk=request.GET.get('house_field'))
        section_house = house.section_set.all().values('id', 'name')
        floor_house = house.floor_set.all().values('id', 'name')
        print(section_house)
        print(floor_house)
        response = {
            'section_data': list(section_house),
            'floor_data': list(floor_house)
        }
        return JsonResponse(response, status=200)


class RequestDetail(DetailView):
    model = Request
    template_name = 'request_detail.html'

@method_decorator(user_passes_test(lambda u: check_user_is_staff(u, 'application'), login_url='login_admin'), name='dispatch')
class RequestsList(ListView):
    model = Request
    template_name = 'requests_list.html'

    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(**kwargs)
        users = UserProfile.objects.select_related('role').all()
        queryset = Request.objects.select_related('apartment', 'apartment__owner', 'apartment__section', 'master').all()
        if self.request.GET.get('input_number'):
            queryset = queryset.filter(pk__contains=self.request.GET.get('input_number'))
        if self.request.GET.get('filter-number') == '1':
            queryset = queryset.order_by('-pk')
        if self.request.GET.get('filter-number') == '0':
            queryset = queryset.order_by('pk')
        if self.request.GET.get('input_date'):
            input_date = self.request.GET.get('input_date').split(' - ')
            date_range = [datetime.datetime.strptime(x, '%d.%m.%Y').strftime('%Y-%m-%d') for x in input_date]
            queryset = queryset.filter(date__range=date_range)
        if self.request.GET.get('filter-date') == '1':
            queryset = queryset.order_by('date')
        if self.request.GET.get('filter-date') == '0':
            queryset = queryset.order_by('-date')
        if self.request.GET.get('input_type'):
            queryset = queryset.filter(type_master=self.request.GET.get('input_type'))
        if self.request.GET.get('filter-master') == '1':
            queryset = queryset.order_by('type_master')
        if self.request.GET.get('filter-master') == '0':
            queryset = queryset.order_by('-type_master')
        if self.request.GET.get('description'):
            queryset = queryset.filter(description__icontains=self.request.GET.get('description'))
        if self.request.GET.get('apartment'):
            queryset = queryset.filter(apartment__number__contains=self.request.GET.get('apartment'))
        if self.request.GET.get('owner'):
            queryset = queryset.filter(apartment__owner_id=self.request.GET.get('owner'))
        if self.request.GET.get('telephone'):
            queryset = queryset.filter(apartment__owner__telephone__contains=self.request.GET.get('telephone'))
        if self.request.GET.get('master'):
            queryset = queryset.filter(master=self.request.GET.get('master'))
        if self.request.GET.get('status'):
            queryset = queryset.filter(status=self.request.GET.get('status'))

        paginator = Paginator(queryset, 10)  # 3 posts in each page
        page = self.request.GET.get('page')
        try:
            queryset = paginator.page(page)
        except PageNotAnInteger:
            # If page is not an integer deliver the first page
            queryset = paginator.page(1)
        except EmptyPage:
            # If page is out of range deliver last page of results
            queryset = paginator.page(paginator.num_pages)

        context['owners'] = users.filter(is_staff=False)
        context['masters'] = users.filter(Q(role__roles='plumber') | Q(role__roles='electric'))
        context['houses'] = House.objects.prefetch_related('section_set__apartment_set').all()
        context['request_list'] = queryset
        return context


class RequestCreate(CreateView):
    model = Request
    template_name = 'request_create.html'
    form_class = RequestForm
    success_url = reverse_lazy('requests_list')


class RequestUpdate(UpdateView):
    model = Request
    template_name = 'request_update.html'
    form_class = RequestForm
    success_url = reverse_lazy('requests_list')


def request_delete(request, pk):
    req = get_object_or_404(Request, id=pk)
    req.delete()

    if request.GET.keys():
        return redirect('owner_request_list')
    else:
        return redirect('requests_list')


def select_apart(request):
    if request.GET.get('user_id'):
        print(request.GET.get('user_id'))
        user_apart = Apartment.objects.filter(owner=request.GET.get('user_id')).values('id', 'number')
        print(user_apart)
        data = {
            'user_apart': list(user_apart),
        }
        return JsonResponse(data, status=200)

@method_decorator(user_passes_test(lambda u: check_user_is_staff(u, 'message'), login_url='login_admin'), name='dispatch')
class MessagesList(ListView):
    model = Message
    template_name = 'messages_list.html'


class MessageCreate(CreateView):
    model = Message
    template_name = 'message_create.html'
    form_class = MessageForm
    success_url = reverse_lazy('messages_list')


class MessageDetail(DetailView):
    model = Message
    template_name = 'message_detail.html'


def message_delete(request, pk):
    message = get_object_or_404(Message, id=pk)
    message.delete()
    if request.GET.keys():
        return redirect('owner_message_list')
    else:
        return redirect('messages_list')


def message_select_house(request):
    if request.GET.get('id_house'):
        house = House.objects.get(pk=request.GET.get('id_house'))
        section_house = house.section_set.all().values('id', 'name')
        floor_house = house.floor_set.all().values('id', 'name')
        apart_house = house.apartment_set.all().values('id', 'number')

        response = {
            'section_data': list(section_house),
            'floor_data': list(floor_house),
            'apart_data': list(apart_house)
        }
        return JsonResponse(response, status=200)


def delete_selected_messages(request):
    print(request.GET)
    list_k = list(request.GET.keys())
    print(list_k)
    if request.GET.keys():
        for idk in list_k[0].split(','):
            message = get_object_or_404(Message, id=idk)
            message.delete()
    if request.GET.get('id'):
        id_message = request.GET.get('id')
        for i in id_message.split(' '):
            message = get_object_or_404(Message, id=i)
            message.delete()
    return JsonResponse({}, status=200)

@method_decorator(user_passes_test(lambda u: check_user_is_staff(u, 'personal_account'), login_url='login_admin'), name='dispatch')
class AccountsList(ListView):
    model = Account
    template_name = 'accounts_list.html'

    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(**kwargs)
        object_list = Account.objects.all()
        if self.request.GET.get('number'):
            object_list = object_list.filter(number__contains=self.request.GET.get('number'))
        if self.request.GET.get('status'):
            object_list = object_list.filter(status=self.request.GET.get('status'))
        if self.request.GET.get('apartment'):
            object_list = object_list.filter(apartment=self.request.GET.get('apartment'))
        if self.request.GET.get('house'):
            sections = get_object_or_404(House, pk=self.request.GET.get('house')).section_set.all()
            object_list = object_list.filter(apartment__section__in=sections)
            context['sections'] = sections

        if self.request.GET.get('section'):
            object_list = object_list.filter(apartment__section__name=self.request.GET.get('section'))
        if self.request.GET.get('owner'):
            object_list = object_list.filter(apartment__owner_id=self.request.GET.get('owner'))
        if self.request.GET.get('debt'):
            if self.request.GET.get('debt') == 'yes':
                object_list = object_list.filter(balance__lt=0)
            else:
                object_list = object_list.filter(balance__gte=0)
        paginator = Paginator(object_list, 5)
        page = self.request.GET.get('page')
        try:
            object_list = paginator.page(page)
        except PageNotAnInteger:
            # If page is not an integer deliver the first page
            object_list = paginator.page(1)
        except EmptyPage:
            # If page is out of range deliver last page of results
            object_list = paginator.page(paginator.num_pages)

        context['object_list'] = object_list
        context['page'] = page
        context['houses'] = House.objects.all()
        context['owners'] = UserProfile.objects.filter(role_id__isnull=True)
        return context


class AccountCreate(CreateView):
    model = Account
    template_name = 'account_create.html'
    form_class = AccountForm
    success_url = reverse_lazy('accounts_list')

    def post(self, request):
        form = self.form_class(request.POST)

        if form.is_valid():
            # Делаем операции с формой
            # Например сохраняем данные в модель
            form.save()
            return HttpResponseRedirect(self.success_url)
        else:
            # Выводим только те ошибки, которые присутствуют в форме
            print(form.errors)
            return render(request, self.template_name, {'form': form})


class AccountUpdate(UpdateView):
    model = Account
    template_name = 'account_update.html'
    form_class = AccountForm
    success_url = reverse_lazy('accounts_list')


class AccountDetail(DetailView):
    model = Account
    template_name = 'account_detail.html'


def account_delete(request, pk):
    try:
        account = get_object_or_404(Account, id=pk)
        account.delete()
        return redirect('accounts_list')
    except:
        return redirect('accounts_list')


def select_account(request):
    if request.GET.get('house_field'):
        house = House.objects.get(pk=request.GET.get('house_field'))
        section_house = house.section_set.all().values('id', 'name')
        if request.META.get('HTTP_REFERER').split('/')[4] == 'invoice':
            apartment_house = house.apartment_set.values('id', 'number')
        else:
            apartment_house = house.apartment_set.filter(account__isnull=True).values('id', 'number')
        response = {
            'section_data': list(section_house),
            'apartment_data': list(apartment_house),
        }
        return JsonResponse(response, status=200)
    if request.GET.get('account_field'):
        owner = UserProfile.objects.get(apartment__account=request.GET.get('account_field'))
        account = Account.objects.get(id=request.GET.get('account_field'))

        owner_house = {
            'house': account.apartment.house.name,
            'house_id': account.apartment.house.id,
            'section': account.apartment.section.name,
            'section_id': account.apartment.section.id,
            'apartment': account.apartment.number,
            'apartment_id': account.apartment.id,
            'id': owner.id,
            'first': owner.first_name,
            'last': owner.last_name,
            'telephone': owner.telephone,
            'account': account.number,
            'account_id': account.id

        }
        response = {
            'owners_data': owner_house,
        }
        return JsonResponse(response, status=200)
    if request.GET.get('section_field'):
        section = Section.objects.get(pk=request.GET.get('section_field'))
        if request.META.get('HTTP_REFERER').split('/')[4] == 'invoice':
            apartment_house = section.apartment_set.values('id', 'number')
        else:
            apartment_house = section.apartment_set.filter(account__isnull=True).values('id', 'number')

        response = {
            'apartment_data': list(apartment_house),
        }
        return JsonResponse(response, status=200)
    if request.GET.get('apartment_field'):
        owner = UserProfile.objects.get(apartment=request.GET.get('apartment_field'))
        meters = MeterReading.objects.filter(apartment=request.GET.get('apartment_field'))
        if request.META.get('HTTP_REFERER').split('/')[4] == 'invoice':
            try:
                account = Account.objects.get(apartment=request.GET.get('apartment_field'))
                tariff = Tariff.objects.get(apartment=request.GET.get('apartment_field'))
            except ObjectDoesNotExist:
                account = None
                tariff = None

        account = get_object_or_404(Account, apartment=request.GET.get('apartment_field'))
        tariff = get_object_or_404(Tariff, apartment=request.GET.get('apartment_field'))


        owner_house = {
            'id': owner.id,
            'first': owner.first_name,
            'last': owner.last_name,
            'telephone': owner.telephone,
            'account': account.number if account else '',
            'account_id': account.id if account else '',
            'tariff_id': tariff.id if tariff else '',

        }
        response = {
            'owners_data': owner_house,
            'meters': list(meters)
        }
        return JsonResponse(response, status=200)

@method_decorator(user_passes_test(lambda u: check_user_is_staff(u, 'invoice'), login_url='login_admin'), name='dispatch')
class InvoicesList(ListView):
    model = Invoice
    template_name = 'invoices_list.html'


class InvoiceCreate(CreateView):
    model = Invoice
    template_name = 'invoice_create.html'
    form_class = InvoiceForm
    success_url = reverse_lazy('invoices_list')

    def get_context_data(self, *args, **kwargs):
        meters = MeterReading.objects.all()
        context = super().get_context_data(**kwargs)
        if self.kwargs.get('pk') != None:
            inst = get_object_or_404(Invoice, id=self.kwargs.get('pk'))
            next_number = Invoice.objects.last().number + 1
            inst.pk = None
            inst.number = next_number
            qs_invoice = InvoiceService.objects.filter(invoice_id=self.kwargs.get('pk'))

            context['form'] = InvoiceForm(instance=inst)
            formset = InvoiceServiceFormSet(self.request.POST or None, prefix='formset',
                                  queryset=qs_invoice)
            formset.management_form.initial['INITIAL_FORMS'] = 0
            context['formset'] = formset
        else:
            context['form'] = InvoiceForm()
            context['formset'] = InvoiceServiceFormSet(self.request.POST or None, prefix='formset',
                                                       queryset=InvoiceService.objects.none())
        context['meters'] = meters
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        formset = context['formset']
        print(self.request.POST)


        if formset.is_valid():
            formset.save(commit=False)
            print('1')
            try:
                account = get_object_or_404(Account, number=self.request.POST.get('account'))
                account.balance = account.balance - int(self.request.POST.get('amount'))
                account.save()
            except:
                pass
            form.save()
            for fo in formset:
                if not fo.cleaned_data.get('DELETE'):
                    item = fo.save(commit=False)
                    item.invoice = form.instance
                    try:
                        item.save()
                    except:
                        messages.error(self.request, "Invalid form")
                        print(formset.errors)
                        return self.render_to_response(self.get_context_data(form=form))

            return super().form_valid(form)
        else:
            print('22222222222222222222222222222222222222222222')
            print(form.errors)
            print(formset.errors)
            messages.error(self.request, "Invalid form")
            return self.render_to_response(self.get_context_data(form=form))

    def form_invalid(self, form):

        print(form.errors)
        return self.render_to_response(self.get_context_data(form=form))

class InvoiceUpdate(UpdateView):
    model = Invoice
    template_name = 'invoice_update.html'
    form_class = InvoiceForm
    success_url = reverse_lazy('invoices_list')

    def get_context_data(self, *args, **kwargs):
        meters = MeterReading.objects.all()
        context = super().get_context_data(**kwargs)
        context['formset'] = InvoiceServiceFormSet(self.request.POST or None, prefix='formset',
                                                   queryset=InvoiceService.objects.filter(invoice=self.get_object().pk))
        context['meters'] = meters
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        formset = context['formset']

        if formset.is_valid():
            print(self.request.POST)
            formset.save(commit=False)
            for obj in formset.deleted_objects:
                obj.delete()

            invoice = self.get_object()
            previous_amount = invoice.amount
            form.save()
            try:
                account = get_object_or_404(Account, number=self.request.POST.get('account'))
                old_balance = account.balance
                new_balance = old_balance + previous_amount - int(self.request.POST.get('amount'))
                account.balance = new_balance
                account.save()
            except:
                pass
            for fo in formset:
                if not fo.cleaned_data.get('DELETE'):
                    item = fo.save(commit=False)
                    item.invoice = form.instance
                    try:
                        item.save()
                    except IntegrityError:
                        messages.error(self.request, "Вы пробуете записать две одинаковые услуги, это запрещено.")
                        return self.form_invalid(form)
            return super().form_valid(form)
        else:
            print(self.request.POST)
            print(formset.errors)
            return self.render_to_response(self.get_context_data(form=form))

class InvoiceDetail(DetailView):
    model = Invoice
    template_name = 'invoice_detail.html'

    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(**kwargs)
        invoice = Invoice.objects.get(id=self.get_object().id)
        invoice_service = InvoiceService.objects.filter(invoice_id=self.get_object().id)
        context['invoice_service'] = invoice_service
        context['invoice'] = invoice
        return context


def invoice_delete(request, pk):
    try:
        invoice_service = get_object_or_404(InvoiceService, id=pk)
        invoice_service.delete()
    except:
        pass
    try:
        account = Account.objects.get(apartment__invoice=pk)
        invoice = get_object_or_404(Invoice, id=pk)
        account.balance = account.balance + invoice.amount
        account.save()
        invoice.delete()
    except:

        invoice = get_object_or_404(Invoice, id=pk)
        invoice.delete()
    return redirect('invoices_list')


def select_invoices(request):
    print(request.GET)
    draw = request.GET['draw']

    invoices = Invoice.objects.all()
    invoices_list = []
    filters = Q()

    if request.GET['filterNumber']:
        filters &= Q(number=request.GET['filterNumber'])
    if request.GET['filterDateDay']:
        filters &= Q(date=request.GET['filterDateDay'])
    if request.GET['filterDateMonth']:
        print(request.GET['filterDateMonth'])
        filters &= Q(date__month=request.GET['filterDateMonth'].split('-')[-1])
    if request.GET['filterStatus']:
        filters &= Q(status=request.GET['filterStatus'])
    if request.GET['filterApartment']:
        filters &= Q(apartment__number=request.GET['filterApartment'])
    if request.GET['filterOwner']:
        filters &= Q(apartment__owner=request.GET['filterOwner'])
    if request.GET['filterDone']:
        filters &= Q(done=request.GET['filterDone'])
    if request.GET['filterAmount']:
        filters &= Q(amount=request.GET['filterAmount'])

    if request.GET['sortDate'] == '1':
        invoices = invoices.filter(filters).order_by('date')
    elif request.GET['sortDate'] == '0':
        invoices = invoices.filter(filters).order_by('-date')


    for invoice in invoices:
        # print(invoice.number)
        invoice_dict = {
            'number': invoice.number,
            'day': invoice.date.strftime('%d.%m.%Y'),
            'month': invoice.date.strftime('%m.%Y'),
            'status': invoice.get_status_display(),
            'apartment': invoice.apartment.__str__(),
            'owner': invoice.apartment.owner.__str__(),
            'done': 'Проведена' if invoice.done else 'Не проведена',
            'amount': invoice.amount,
            'id': invoice.id
        }
        invoices_list.append(invoice_dict)

    page_number = request.GET.get('page')
    paginator = Paginator(invoices_list, 10)
    page_obj = paginator.get_page(page_number)
    response = {
        'draw': draw,
        'recordsTotal': paginator.count,
        'recordsFiltered': paginator.count,
        'data': list(page_obj.object_list),
    }
    return JsonResponse(response, status=200)


def delete_selected_invoice(request):
    list_k = list(request.GET.keys())
    if request.GET.keys():
        for idk in list_k[0].split(','):
            if idk != '':
                invoice = get_object_or_404(Invoice, id=idk)
                invoice.delete()

    if request.GET.get('id'):
        id_invoice = request.GET.get('id')
        for i in id_invoice.split(' '):
            invoice = get_object_or_404(Invoice, id=i)
            invoice.delete()
    return JsonResponse({}, status=200)


def invoice_unit(request):
    if request.GET.get('service_id'):
        unit = Unit.objects.get(services=request.GET.get('service_id'))

        invoice_unit = {
            'unit_id': unit.id,
            'unit': unit.name,
        }
        response = {
            'unit_data': invoice_unit,

        }
        return JsonResponse(response, status=200)

@method_decorator(user_passes_test(lambda u: check_user_is_staff(u, 'cashbox'), login_url='login_admin'), name='dispatch')
class TransfersList(ListView):
    model = Transfers
    template_name = 'transfers_list.html'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['object_list'] = Transfers.objects.all()
        context['owners'] = UserProfile.objects.filter(role_id__isnull=True)
        context['items'] = Item.objects.all()
        context['income'] = Transfers.objects.filter(completed=True, income=True).aggregate(total_amount=Sum('amount'))[
            'total_amount']
        expense = Transfers.objects.filter(completed=True, income=False).aggregate(total_amount=Sum('amount'))[
            'total_amount']
        context['expense'] = abs(expense) if expense else '0'

        return context


class TransferCreate(CreateView):
    model = Transfers
    template_name = 'transfer_create_income.html'
    form_class = TransferForm
    success_url = reverse_lazy('transfers_list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(**kwargs)
        transfer_type = self.request.GET.get('type')
        apart_account = self.request.GET.get('account')
        if self.kwargs.get('pk') != None:
            idx = int(Transfers.objects.all().last().number) + 1
            inst = get_object_or_404(Transfers, id=self.kwargs.get('pk'))
            context['form'] = TransferForm(self.request.user, instance=inst, initial={'number': idx})
        if transfer_type == 'income' or apart_account != None:
            self.template_name = 'transfer_create_income.html'
        elif transfer_type == 'out' or apart_account != None:
            self.template_name = 'transfer_create_out.html'
        return context

    def form_valid(self, form):

        if form.is_valid():

            if self.request.POST.get('income') == 'True':
                account = get_object_or_404(Account, id=self.request.POST.get('account'))
                account.balance = account.balance + int(self.request.POST.get('amount').split('.')[0])
                account.save()
            else:
                transfer = form.save(commit=False)
                transfer.amount = -int(self.request.POST.get('amount').split('.')[0])
                transfer.save()
            form.save()
            return HttpResponseRedirect(self.success_url)

        else:
            # Выводим только те ошибки, которые присутствуют в форме

            return render(self.request, self.template_name, {'form': form})
    def form_invalid(self, form):
        print(form.errors)

class TransferUpdate(UpdateView):
    model = Transfers
    template_name = 'transfer_update.html'
    form_class = TransferForm
    success_url = reverse_lazy('transfers_list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(**kwargs)
        transfer_type = Transfers.objects.get(id=self.get_object().pk)
        if transfer_type.income:
            self.template_name = 'transfer_update_income.html'
        else:
            self.template_name = 'transfer_update_out.html'
        return context

    def form_valid(self, form):
        print(self.request.POST)
        if self.request.POST.get('income') == 'True':
            try:
                account = get_object_or_404(Account, id=self.request.POST.get('account'))
                transfer = get_object_or_404(Transfers, id=self.get_object().pk)
                old_balance = account.balance - transfer.amount
                account.balance = old_balance + int(self.request.POST.get('amount').split('.')[0])
                account.save()
            except:
                pass

            super(TransferUpdate, self).form_valid(form)
            return redirect('transfer_detail', self.object.id)


class TransferDetail(DetailView):
    model = Transfers
    template_name = 'transfer_detail.html'

    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(**kwargs)
        context['transfers'] = Transfers.objects.get(id=self.get_object().pk)
        # context['invoice'] = Invoice.objects.filter(da=True)
        # context['items'] = Item.objects.all()
        return context


def transfer_delete(request, pk):
    transfer = get_object_or_404(Transfers, id=pk)
    if transfer.income:
        try:
            account = get_object_or_404(Account, transfers=pk)
            account.balance = account.balance - transfer.amount
            account.save()
        except:
            print('xo xo xo')
    transfer.delete()
    return redirect('transfers_list')


def select_transfers(request):
    print(request.GET)
    draw = request.GET['draw']

    transfers = Transfers.objects.all()
    transfers_list = []
    filters = Q()

    if request.GET['filterNumber']:
        filters &= Q(number=request.GET['filterNumber'])
    if request.GET['filterDate']:
        filters &= Q(date=request.GET.get('filterDate'))
    if request.GET['filterCompleted']:
        filters &= Q(completed=request.GET['filterCompleted'])
    if request.GET['filterItem']:
        filters &= Q(item=request.GET['filterItem'])
    if request.GET['filterOwner']:
        filters &= Q(owner=request.GET['filterOwner'])
    if request.GET['filterAccount']:
        filters &= Q(account__number=request.GET['filterAccount'])
    if request.GET['filterIncome']:
        filters &= Q(income=request.GET['filterIncome'])

    if request.GET['sortDate'] == '1':
        transfers = transfers.filter(filters).order_by('date')
    elif request.GET['sortDate'] == '0':
        transfers = transfers.filter(filters).order_by('-date')
    for trans in transfers:
        # print(invoice.number)
        transfer_dict = {
            'number': trans.number,
            'date': trans.date.strftime('%d.%m.%Y'),
            'completed': 'Проведена' if trans.completed else 'Не проведена',
            'item': trans.item.name,
            'owner': trans.owner.__str__() if trans.owner else 'Не заданно',
            'account': trans.account.number if trans.account else 'Не заданно',
            'income': 'Приход' if trans.income else 'Расход',
            'amount': trans.amount,
            'id': trans.id
        }
        transfers_list.append(transfer_dict)

    page_number = request.GET.get('page')
    paginator = Paginator(transfers_list, 10)
    page_obj = paginator.get_page(page_number)

    response = {
        'draw': draw,
        'recordsTotal': paginator.count,
        'recordsFiltered': paginator.count,
        'data': list(page_obj.object_list),
    }
    return JsonResponse(response, status=200)


def select_owners(request):
    if request.GET.get('owner_field'):
        accounts = Account.objects.filter(apartment__owner_id=request.GET.get('owner_field')).values('id', 'number')
        response = {
            'accounts_data': list(accounts),
        }
        return JsonResponse(response, status=200)


def tariff_set(request):
    tariffs = PriceTariffServices.objects.filter(tariff_id=request.GET.get('tariff_id'))

    print(request.META['HTTP_REFERER'].split('/')[5])
    if request.META['HTTP_REFERER'].split('/')[5] == 'update':
        invserdel = InvoiceService.objects.filter(invoice_id=request.META['HTTP_REFERER'].split('/')[6])
        for ins in invserdel:
            ins.delete()
    tariff_list = []
    for tariff in tariffs:
        print(tariff.tariff.id)
        tariff_dict = {

            'price': tariff.price,
            'id': tariff.services.id,
            'unit': tariff.services.unit.name,
        }
        tariff_list.append(tariff_dict)
    return JsonResponse({'data': tariff_list}, status=200)


class TemplatesList(ListView):
    model = Template
    template_name = 'templates_list.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        try:
            default_template = Template.objects.get(type=True)
        except:
            default_template = Template.objects.create(name='test', type=True, file='/media/files/example.xlsx')
        context['object_list'] = Template.objects.all()
        context['template_default'] = default_template
        context['invoice'] = get_object_or_404(Invoice, id=self.request.GET.get('invoice'))
        return context


class TemplateConfig(CreateView):
    model = Template
    template_name = 'template_config.html'
    form_class = TemplateForm
    success_url = reverse_lazy('template_config')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['object_list'] = Template.objects.all()
        return context


def set_default(request):
    if request.method == 'GET':
        print(request.GET.get('template_id'))
        old_template = Template.objects.get(type=True)
        old_template.type = False
        old_template.save()
        template = get_object_or_404(Template, pk=request.GET.get('template_id'))
        template.type = True
        template.save()
        return JsonResponse({}, status=200)


def template_delete(request):
    if request.method == 'GET':
        print(request.GET.get('template_id'))
        templates = Template.objects.all()
        template = templates.get(pk=request.GET.get('template_id'))
        if templates.count() > 1:
            if template.type == True:
                template_new = templates.exclude(pk=request.GET.get('template_id')).first()
                template_new.type = True
                template_new.save()
                template.delete()
                return JsonResponse({'data': template_new.id}, status=200)
            else:
                template.delete()
                return JsonResponse({}, status=200)
        else:
            return JsonResponse({}, status=400)


def template_upload(request):
    template = Template.objects.get(id=request.GET.get('idx'))
    invoice = Invoice.objects.get(id=request.GET.get('invoice_id'))
    invoice_service = InvoiceService.objects.filter(invoice_id=request.GET.get('invoice_id'))
    wb = load_workbook(template.file)

    sheet = wb.active

    # Поиск и замена значения
    search_word = ['%invoiceAddress%',
                   '%payCompany%',
                   '%accountNumber%',
                   '%invoiceNumber%',
                   '%invoiceDate%',
                   '%accountBalance%',
                   '%totalDebt%',
                   '%invoiceDate%',
                   '%invoiceMonth%',
                   '%total%']
    try:
        pay = invoice.apartment.account.balance - invoice.amount if invoice.apartment.account.balance < invoice.amount else '0.00'
        number = invoice.apartment.account.number
        balance = invoice.apartment.account.balance
    except:
        pay = 'Не заданно'
        number = 'Не заданно'
        balance = 'He заданно'

    locale.setlocale(locale.LC_TIME, 'ru_RU.UTF-8')
    if invoice.apartment.house.address:
        address = invoice.apartment.house.address
    else:
        address = "Не указан"
    replacement = [address,
                   'moidom24',
                   number,
                   invoice.number,
                   invoice.date.strftime('%d.%m.%Y'),
                   balance,
                   pay[0],
                   invoice.date.strftime('%d.%m.%Y'),
                   invoice.date.strftime('%B'),
                   invoice.amount]

    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    font = Font(name='Arial', size=14, bold=True)

    for row in sheet.iter_rows():
        for cell in row:

            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            for search in search_word:
                if cell.value == search:
                    cell.value = replacement[int(search_word.index(search))]

    row = 19
    for service in invoice_service:
        sheet.cell(row=row, column=1).value = service.service.main
        sheet.cell(row=row, column=3).value = service.cost_for_unit
        sheet.cell(row=row, column=5).value = service.service.unit.name
        sheet.cell(row=row, column=7).value = service.expense
        sheet.cell(row=row, column=9).value = service.full_cost
        row += 1

    sheet.cell(row=row, column=7).value = "РАЗОМ:"
    sheet.cell(row=row, column=9).value = float(invoice.amount)
    for column in range(1, 12):
        sheet.cell(row=row, column=column).border = border
        sheet.cell(row=row, column=column).font = font


    # Сохранение изменений в файл
    wb.save('media/files/example.xlsx')
    if request.GET.get('send') == '1':
        return JsonResponse({'data': '/media/files/example.xlsx'}, status=200)
    elif request.GET.get('send') == '0' and invoice.apartment.owner.email:
        email = EmailMessage(
            'Вот вам квитанция',
            'Поздравлямс',
            None,
            [invoice.apartment.owner.email]
        )
        email.attach_file('media/files/example.xlsx')
        email.send()
        return JsonResponse({}, status=200)
    elif request.GET.get('send') == '2':

        workbook = load_workbook(filename='media/files/example.xlsx')
        worksheet = workbook.active
        pdfmetrics.registerFont(TTFont('Arial', 'media/fonts/ArialMT.ttf'))

        data = []
        for row in worksheet.iter_rows(min_row=0, values_only=True):
            data.append(row[:-1])
        doc = SimpleDocTemplate('media/files/example.pdf', pagesize=landscape(A4))
        doc.leftMargin = doc.rightMargin = doc.topMargin = doc.bottomMargin = 36
        elements = []

        # Define table style
        style = TableStyle([

            ('BACKGROUND', (0, 0), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), 'Arial'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),
            ('BOX', (0, 0), (-1, -1), 0.25, colors.black),



            ('SPAN', (1, 0), (6, 2)),
            ('VALIGN', (1, 0), (6, 2), 'MIDDLE'),
            ('SPAN', (1, 4), (6, 4)),
            ('VALIGN', (1, 4), (6, 4), 'MIDDLE'),
            ('SPAN', (8, 0), (9, 0)),
            ('VALIGN', (8, 0), (9, 0), 'MIDDLE'),
            ('SPAN', (0, 8), (3, 8)),
            ('VALIGN', (0, 8), (3, 8), 'MIDDLE'),
            ('SPAN', (5, 7), (7, 8)),
            ('VALIGN', (5, 7), (7, 8), 'MIDDLE'),
            ('SPAN', (1, 9), (6, 11)),
            ('VALIGN', (1, 9), (6, 11), 'MIDDLE'),
            ('SPAN', (8, 9), (9, 9)),
            ('VALIGN', (8, 9), (9, 9), 'MIDDLE'),

            ('VALIGN', (1, 13), (6, 13), 'MIDDLE'),
            ('SPAN', (1, 13), (6, 13)),
            ('VALIGN', (0, 17), (1, 17), 'MIDDLE'),
            ('SPAN', (0, 17), (1, 17)),
            ('VALIGN', (2, 17), (3, 17), 'MIDDLE'),
            ('SPAN', (2, 17), (3, 17)),
            ('VALIGN', (4, 17), (5, 17), 'MIDDLE'),
            ('SPAN', (4, 17), (5, 17)),
            ('VALIGN', (6, 17), (7, 17), 'MIDDLE'),
            ('SPAN', (6, 17), (7, 17)),

        ])
        # Create table and apply style

        table = Table(data, colWidths=70)

        table.setStyle(style)
        elements.append(table)

        # Build PDF document
        doc.build(elements)

        # # Создаем PDF-файл
        # pdf_file = canvas.Canvas('media/files/example.pdf', pagesize=landscape(A4))
        # pdf_file.setFont('Arial', 8)
        # x = 70
        # y = 500
        # table_width = len(data[0]) * x
        # table_height = 22 * 15
        # # Рисуем рамку вокруг таблицы
        # pdf_file.rect(30-2, y - table_height + 13, table_width, table_height)
        # pdf_file.setStrokeColor(colors.lightgrey)
        # for row in data:
        #     x = 30
        #     for item in row:
        #         if not item:
        #             pdf_file.drawString(x, y, '')
        #             pdf_file.rect(x - 2, y - 2, 70, 15, stroke=True)
        #             x += 70
        #
        #         else:
        #             if '\n' in str(item):
        #                 pdf_file.drawString(x, y, str(item).split('\n')[0])
        #                 y -= 15
        #                 pdf_file.drawString(x, y, str(item).split('\n')[1])
        #                 y += 15
        #                 x += 70
        #             else:
        #                 pdf_file.drawString(x, y, str(item))
        #                 pdf_file.rect(x - 2, y - 2, 70, 15)
        #                 x += 70
        #
        #     y -= 15
        #
        # print(data)
        # pdf_file.save()

        return JsonResponse({'data': '/media/files/example.pdf'}, status=200)
    else:
        return JsonResponse({}, status=400)
